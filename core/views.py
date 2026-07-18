import io
import json
import base64
import hashlib
import secrets
import string
import subprocess
import sys
import zipfile
from decimal import Decimal
import requests
from django.core.files.base import ContentFile
from django.db import models
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.conf import settings
from django.utils import timezone
from rest_framework import viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from .models import (
    Bookmark, ClipboardItem, Expense, FileConversion, GeneratedPassword,
    ImageAsset, ImageProcessHistory, IPLookupHistory, Note, Reminder, SavedQRCode,
    ShortLink, Todo, ToolUsage,
)
from .serializers import (
    BookmarkSerializer, ClipboardItemSerializer, ExpenseSerializer,
    GeneratedPasswordSerializer, ImageAssetSerializer, ImageProcessHistorySerializer,
    IPLookupHistorySerializer, NoteSerializer, ReminderSerializer,
    SavedQRCodeSerializer, ShortLinkSerializer, TodoSerializer,
    ToolUsageSerializer, UploadedFileSerializer,
)

def index(request):
    index_path = settings.BASE_DIR / 'static' / 'vue' / 'index.html'
    if index_path.exists():
        with open(index_path, 'r', encoding='utf-8') as f:
            return HttpResponse(f.read(), content_type='text/html')
    return HttpResponse('Frontend is building...', content_type='text/html')

def api_root(request):
    return JsonResponse({
        "name": "ToolHub API",
        "version": "1.0",
        "endpoints": {
            "auth": {"register": "/api/auth/register/", "login": "/api/auth/login/", "me": "/api/auth/me/", "logout": "/api/auth/logout/"},
            "notes": "/api/notes/", "todos": "/api/todos/", "passwords": "/api/passwords/",
            "qrcodes": "/api/qrcodes/", "ip_lookups": "/api/ip-lookups/", "files": "/api/files/",
            "tools": {"qrcode": "/api/qrcode/?text=...", "ip_lookup": "/api/ip-lookup/?q=..."},
        },
    })


def _generate_short_code(length=6):
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


class NoteViewSet(viewsets.ModelViewSet):
    queryset = Note.objects.none()
    serializer_class = NoteSerializer
    def get_queryset(self):
        return Note.objects.filter(user=self.request.user)
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class TodoViewSet(viewsets.ModelViewSet):
    queryset = Todo.objects.none()
    serializer_class = TodoSerializer
    def get_queryset(self):
        return Todo.objects.filter(user=self.request.user)
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class GeneratedPasswordViewSet(viewsets.ModelViewSet):
    queryset = GeneratedPassword.objects.none()
    serializer_class = GeneratedPasswordSerializer
    def get_queryset(self):
        return GeneratedPassword.objects.filter(user=self.request.user)
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class SavedQRCodeViewSet(viewsets.ModelViewSet):
    queryset = SavedQRCode.objects.none()
    serializer_class = SavedQRCodeSerializer
    def get_queryset(self):
        return SavedQRCode.objects.filter(user=self.request.user)
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class IPLookupHistoryViewSet(viewsets.ModelViewSet):
    queryset = IPLookupHistory.objects.none()
    serializer_class = IPLookupHistorySerializer
    def get_queryset(self):
        return IPLookupHistory.objects.filter(user=self.request.user)
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

class ToolUsageViewSet(viewsets.ModelViewSet):
    queryset = ToolUsage.objects.none()
    serializer_class = ToolUsageSerializer
    def get_queryset(self):
        return ToolUsage.objects.filter(user=self.request.user)
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ClipboardItemViewSet(viewsets.ModelViewSet):
    queryset = ClipboardItem.objects.none()
    serializer_class = ClipboardItemSerializer

    def get_queryset(self):
        qs = ClipboardItem.objects.filter(user=self.request.user)
        q = self.request.query_params.get("q", "").strip()
        if q:
            qs = qs.filter(models.Q(title__icontains=q) | models.Q(content__icontains=q) | models.Q(tags__icontains=q))
        return qs

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
        ToolUsage.objects.create(user=self.request.user, tool="clipboard", action="create")


class BookmarkViewSet(viewsets.ModelViewSet):
    queryset = Bookmark.objects.none()
    serializer_class = BookmarkSerializer

    def get_queryset(self):
        qs = Bookmark.objects.filter(user=self.request.user)
        q = self.request.query_params.get("q", "").strip()
        if q:
            qs = qs.filter(models.Q(title__icontains=q) | models.Q(url__icontains=q) | models.Q(category__icontains=q))
        return qs

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
        ToolUsage.objects.create(user=self.request.user, tool="bookmark", action="create")

    @action(detail=True, methods=["post"])
    def opened(self, request, pk=None):
        item = self.get_object()
        item.open_count += 1
        item.save(update_fields=["open_count"])
        return JsonResponse({"open_count": item.open_count})


class ReminderViewSet(viewsets.ModelViewSet):
    queryset = Reminder.objects.none()
    serializer_class = ReminderSerializer

    def get_queryset(self):
        return Reminder.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
        ToolUsage.objects.create(user=self.request.user, tool="reminder", action="create")


class ShortLinkViewSet(viewsets.ModelViewSet):
    queryset = ShortLink.objects.none()
    serializer_class = ShortLinkSerializer

    def get_queryset(self):
        return ShortLink.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        code = serializer.validated_data.get("code") or _generate_short_code()
        while ShortLink.objects.filter(code=code).exists():
            code = _generate_short_code()
        serializer.save(user=self.request.user, code=code)
        ToolUsage.objects.create(user=self.request.user, tool="shortlink", action="create")


class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.none()
    serializer_class = ExpenseSerializer

    def get_queryset(self):
        return Expense.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
        ToolUsage.objects.create(user=self.request.user, tool="expense", action="create")


class ImageProcessHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ImageProcessHistory.objects.none()
    serializer_class = ImageProcessHistorySerializer

    def get_queryset(self):
        return ImageProcessHistory.objects.filter(user=self.request.user)


class ImageAssetViewSet(viewsets.ModelViewSet):
    queryset = ImageAsset.objects.none()
    serializer_class = ImageAssetSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        qs = ImageAsset.objects.filter(user=self.request.user, is_user_deleted=False)
        q = self.request.query_params.get("q", "").strip()
        category = self.request.query_params.get("category", "").strip()
        favorite = self.request.query_params.get("favorite", "").strip()
        duplicates = self.request.query_params.get("duplicates", "").strip()
        if q:
            qs = qs.filter(
                models.Q(original_filename__icontains=q)
                | models.Q(title__icontains=q)
                | models.Q(tags__icontains=q)
            )
        if category:
            qs = qs.filter(category=category)
        if favorite == "1":
            qs = qs.filter(is_favorite=True)
        if duplicates == "1":
            duplicate_hashes = (
                ImageAsset.objects.filter(user=self.request.user)
                .filter(is_user_deleted=False)
                .exclude(content_hash="")
                .values("content_hash")
                .annotate(total=models.Count("id"))
                .filter(total__gt=1)
                .values_list("content_hash", flat=True)
            )
            qs = qs.filter(content_hash__in=duplicate_hashes)
        return qs

    def perform_create(self, serializer):
        upload = self.request.FILES.get("image")
        metadata = {}
        if upload:
            raw = upload.read()
            upload.seek(0)
            metadata["original_filename"] = upload.name
            metadata["file_size"] = upload.size
            metadata["content_hash"] = hashlib.sha256(raw).hexdigest()
            try:
                from PIL import Image
                img = Image.open(io.BytesIO(raw))
                metadata["width"] = img.width
                metadata["height"] = img.height
            except Exception:
                pass
        serializer.save(user=self.request.user, **metadata)
        ToolUsage.objects.create(user=self.request.user, tool="image", action="organize-upload")

    @action(detail=True, methods=["post"])
    def favorite(self, request, pk=None):
        image = self.get_object()
        image.is_favorite = not image.is_favorite
        image.save(update_fields=["is_favorite", "updated_at"])
        return JsonResponse({"is_favorite": image.is_favorite})

    @action(detail=False, methods=["get"])
    def categories(self, request):
        data = (
            ImageAsset.objects.filter(user=request.user)
            .exclude(category="")
            .values_list("category", flat=True)
            .distinct()
            .order_by("category")
        )
        return JsonResponse({"categories": list(data)})

    @action(detail=False, methods=["post"])
    def batch_compress(self, request):
        ids = request.data.get("ids", [])
        quality = max(10, min(95, int(request.data.get("quality") or 80)))
        fmt = request.data.get("format", "WEBP").upper()
        if fmt not in ("WEBP", "JPEG", "PNG"):
            fmt = "WEBP"
        ext = "jpg" if fmt == "JPEG" else fmt.lower()
        qs = ImageAsset.objects.filter(user=request.user, id__in=ids, is_user_deleted=False)
        if not qs.exists():
            return JsonResponse({"error": "请选择图片"}, status=400)

        zip_buffer = io.BytesIO()
        from PIL import Image
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for asset in qs:
                img = Image.open(asset.image.path)
                if img.mode not in ("RGB", "RGBA"):
                    img = img.convert("RGB")
                output = io.BytesIO()
                kwargs = {"format": fmt}
                if fmt in ("WEBP", "JPEG"):
                    kwargs["quality"] = quality
                img.save(output, **kwargs)
                base = _os.path.splitext(asset.original_filename or f"image-{asset.id}")[0]
                zf.writestr(f"{base}.{ext}", output.getvalue())
        zip_buffer.seek(0)
        ToolUsage.objects.create(user=request.user, tool="image", action="batch-compress", detail=f"{qs.count()} images")
        response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = 'attachment; filename="compressed-images.zip"'
        return response

    @action(detail=True, methods=["post"])
    def compress_to_target(self, request, pk=None):
        asset = self.get_object()
        target_kb = int(request.data.get("target_kb") or 1024)
        fmt = request.data.get("format", "WEBP").upper()
        data, ext, final_size, final_width, final_height = _compress_image_file_to_target(
            asset.image.path,
            target_kb=target_kb,
            fmt=fmt,
        )
        base = _os.path.splitext(asset.original_filename or f"image-{asset.id}")[0]
        ToolUsage.objects.create(
            user=request.user,
            tool="image",
            action="target-compress",
            detail=f"{asset.original_filename} <= {target_kb}KB",
        )
        response = HttpResponse(data, content_type=f"image/{'jpeg' if ext == 'jpg' else ext}")
        response["Content-Disposition"] = f'attachment; filename="{base}_under_{target_kb}kb.{ext}"'
        response["X-Compressed-Size"] = str(final_size)
        response["X-Compressed-Width"] = str(final_width)
        response["X-Compressed-Height"] = str(final_height)
        return response

    def perform_destroy(self, instance):
        instance.is_user_deleted = True
        instance.user_deleted_at = timezone.now()
        instance.is_favorite = False
        instance.save(update_fields=["is_user_deleted", "user_deleted_at", "is_favorite", "updated_at"])

def qrcode_api(request):
    try:
        import qrcode
        from PIL import Image
        text = request.GET.get('text', '')
        size = int(request.GET.get('size', 200))
        save = request.GET.get('save', '0')
        if not text:
            return JsonResponse({'error': '\u8bf7\u63d0\u4f9btext \u53c2\u6570'}, status=400)
        qr = qrcode.QRCode(box_size=10, border=2)
        qr.add_data(text)
        qr.make(fit=True)
        img = qr.make_image(fill_color='black', back_color='white')
        img = img.resize((size, size), Image.NEAREST)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        if save == '1':
            user = request.user if request.user.is_authenticated else None
            SavedQRCode.objects.create(user=user, content=text, size=size)
            ToolUsage.objects.create(user=user, tool='qrcode', action='generate', detail=text[:100])
        return HttpResponse(buf.getvalue(), content_type='image/png')
    except ImportError:
        return JsonResponse({'error': 'QRCode library not installed'}, status=500)

def ip_lookup_api(request):
    q = request.GET.get('q', '')
    save = request.GET.get('save', '0')
    if not q:
        return JsonResponse({'error': '\u8bf7\u63d0\u4f9bq \u53c2\u6570'}, status=400)
    try:
        r = requests.get(f'http://ip-api.com/json/{q}', timeout=5)
        data = r.json()
        if data.get('status') == 'success':
            result = {
                'ip': data.get('query', ''),
                'country': data.get('country', ''),
                'region': data.get('regionName', ''),
                'city': data.get('city', ''),
                'org': data.get('isp', '') or data.get('org', ''),
            }
            if save == '1':
                user = request.user if request.user.is_authenticated else None
                IPLookupHistory.objects.create(user=user, query=q, **result)
                ToolUsage.objects.create(user=user, tool='ip', action='lookup', detail=q)
            return JsonResponse(result)
        else:
            return JsonResponse({'error': data.get('message', '\u67e5\u8be2\u5931\u8d25')}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def password_save_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            pwd = data.get('password', '')
            length = data.get('length', 0)
            note = data.get('note', '')
            user = request.user if request.user.is_authenticated else None
            pwd_obj = GeneratedPassword.objects.create(
                user=user, password=pwd, length=length, note=note,
                has_upper=data.get('has_upper', False),
                has_lower=data.get('has_lower', False),
                has_digits=data.get('has_digits', False),
                has_symbols=data.get('has_symbols', False),
            )
            ToolUsage.objects.create(user=user, tool='password', action='generate')
            return JsonResponse({'id': pwd_obj.id, 'password': pwd_obj.password, 'note': pwd_obj.note,
                                 'created_at': pwd_obj.created_at.isoformat()})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'POST required'}, status=405)

import mimetypes
from django.shortcuts import get_object_or_404
from django.http import FileResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.decorators import action
from .models import UploadedFile
from .serializers import UploadedFileSerializer

class UploadedFileViewSet(viewsets.ModelViewSet):
    queryset = UploadedFile.objects.none()
    serializer_class = UploadedFileSerializer
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        return UploadedFile.objects.filter(user=self.request.user, is_user_deleted=False)

    def perform_create(self, serializer):
        file_obj = self.request.FILES.get('file')
        if file_obj:
            serializer.save(
                user=self.request.user,
                original_filename=file_obj.name,
                file_size=file_obj.size,
                file_type=file_obj.content_type or '',
            )
        else:
            serializer.save(user=self.request.user)

    @action(detail=True, methods=['post'])
    def favorite(self, request, pk=None):
        file_obj = self.get_object()
        file_obj.is_favorite = not file_obj.is_favorite
        file_obj.save(update_fields=['is_favorite'])
        return JsonResponse({'is_favorite': file_obj.is_favorite})

    @action(detail=False, methods=['get'])
    def favorites(self, request):
        qs = UploadedFile.objects.filter(user=request.user, is_favorite=True, is_user_deleted=False)
        serializer = self.get_serializer(qs, many=True)
        return JsonResponse(serializer.data, safe=False)

    def perform_destroy(self, instance):
        instance.is_user_deleted = True
        instance.user_deleted_at = timezone.now()
        instance.is_favorite = False
        instance.save(update_fields=["is_user_deleted", "user_deleted_at", "is_favorite"])

from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from django.contrib.auth.models import User
def _get_client_ip(request):
    """获取客户端真实IP"""
    xff = request.META.get("HTTP_X_FORWARDED_FOR")
    if xff:
        return xff.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")




@api_view(['POST'])
@permission_classes([AllowAny])
def register_api(request):
    import json
    try:
        data = json.loads(request.body)
        username = data.get('username', '').strip()
        password = data.get('password', '')
        email = data.get('email', '')
        if not username or not password:
            return JsonResponse({'error': '\u7528\u6237\u540d\u548c\u5bc6\u7801\u4e0d\u80fd\u4e3a\u7a7a'}, status=400)
        if len(password) < 6:
            return JsonResponse({'error': '\u5bc6\u7801\u81f3\u5c11\u97006\u4f4d'}, status=400)
        if User.objects.filter(username=username).exists():
            return JsonResponse({'error': '\u7528\u6237\u540d\u5df2\u5b58\u5728'}, status=400)
        user = User.objects.create_user(username=username, password=password, email=email)
        token, _ = Token.objects.get_or_create(user=user)
        # 记录登录日志
        try:
            from .models import AccessLog
            AccessLog.objects.create(
                log_type="login", user=user,
                ip_address=_get_client_ip(request),
                path="/api/auth/login/", method="POST",
                user_agent=request.META.get("HTTP_USER_AGENT", "")[:500],
                detail=f"用户 {user.username} 登录"
            )
        except:
            pass
        return JsonResponse({'token': token.key, 'username': user.username, 'user_id': user.id, 'is_superuser': user.is_superuser})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@api_view(['POST'])
@permission_classes([AllowAny])
def login_api(request):
    import json
    try:
        data = json.loads(request.body)
        username = data.get('username', '')
        password = data.get('password', '')
        user = authenticate(username=username, password=password)
        if user:
            token, _ = Token.objects.get_or_create(user=user)
            try:
                from .models import AccessLog
                AccessLog.objects.create(
                    log_type="login", user=user,
                    ip_address=_get_client_ip(request),
                    path="/api/auth/login/", method="POST",
                    user_agent=request.META.get("HTTP_USER_AGENT", "")[:500],
                    detail=f"用户 {user.username} 登录"
                )
            except Exception:
                pass
            return JsonResponse({'token': token.key, 'username': user.username, 'user_id': user.id, 'is_superuser': user.is_superuser})
        return JsonResponse({'error': '\u7528\u6237\u540d\u6216\u5bc6\u7801\u9519\u8bef'}, status=401)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@api_view(['GET'])
def me_api(request):
    token_key = request.META.get('HTTP_AUTHORIZATION', '').replace('Token ', '')
    try:
        token = Token.objects.get(key=token_key)
        return JsonResponse({'username': token.user.username, 'user_id': token.user.id, 'is_superuser': token.user.is_superuser})
    except Token.DoesNotExist:
        return JsonResponse({'error': '\u672a\u767b\u5f55'}, status=401)

@api_view(['POST'])
def logout_api(request):
    try:
        request.user.auth_token.delete()
        return JsonResponse({'message': '\u5df2\u9000\u51fa'})
    except:
        return JsonResponse({'message': 'OK'})

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def file_download(request, pk):
    file_obj = get_object_or_404(UploadedFile, pk=pk, user=request.user, is_user_deleted=False)
    file_obj.downloads_count += 1
    file_obj.save(update_fields=['downloads_count'])
    response = FileResponse(file_obj.file, as_attachment=True, filename=file_obj.original_filename)
    return response

def file_share(request, token):
    """分享链接直接下载文件"""
    file_obj = get_object_or_404(UploadedFile, share_token=token, is_user_deleted=False)
    if file_obj.share_expires_at and timezone.now() > file_obj.share_expires_at:
        return JsonResponse({"error": "分享链接已过期"}, status=410)
    if file_obj.share_max_downloads and file_obj.downloads_count >= file_obj.share_max_downloads:
        return JsonResponse({"error": "分享下载次数已用完"}, status=410)
    if file_obj.share_password and request.GET.get("pwd", "") != file_obj.share_password:
        return JsonResponse({"error": "分享密码错误或缺失"}, status=403)
    file_obj.downloads_count += 1
    file_obj.save(update_fields=["downloads_count"])
    response = FileResponse(file_obj.file, as_attachment=True, filename=file_obj.original_filename)
    return response


def _compress_image_file_to_target(path_or_file, target_kb=1024, fmt="WEBP"):
    from PIL import Image

    target_bytes = max(20, int(target_kb)) * 1024
    fmt = (fmt or "WEBP").upper()
    if fmt not in ("WEBP", "JPEG", "PNG"):
        fmt = "WEBP"
    ext = "jpg" if fmt == "JPEG" else fmt.lower()

    img = Image.open(path_or_file)
    if img.mode not in ("RGB", "RGBA"):
        img = img.convert("RGB")
    if fmt == "JPEG" and img.mode == "RGBA":
        bg = Image.new("RGB", img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[-1])
        img = bg

    def render(source, quality):
        output = io.BytesIO()
        kwargs = {"format": fmt, "optimize": True}
        if fmt in ("WEBP", "JPEG"):
            kwargs["quality"] = quality
        source.save(output, **kwargs)
        return output.getvalue()

    best = None
    current = img
    for scale_step in range(0, 9):
        if scale_step:
            scale = 0.92 ** scale_step
            size = (max(1, int(img.width * scale)), max(1, int(img.height * scale)))
            current = img.resize(size, Image.Resampling.LANCZOS)

        qualities = [95, 92, 88, 84, 80, 76, 72, 68, 64, 60, 55, 50, 45, 40]
        if fmt == "PNG":
            qualities = [95]
        for quality in qualities:
            data = render(current, quality)
            if best is None or len(data) < len(best):
                best = data
            if len(data) <= target_bytes:
                return data, ext, len(data), current.width, current.height

    return best, ext, len(best), current.width, current.height


def shortlink_redirect(request, code):
    link = get_object_or_404(ShortLink, code=code)
    link.visits += 1
    link.save(update_fields=["visits"])
    return redirect(link.target_url)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def dashboard_summary_api(request):
    user = request.user
    today = timezone.localdate()
    month_start = today.replace(day=1)
    expense_total = Expense.objects.filter(user=user, spent_at__gte=month_start).aggregate(total=Sum("amount"))["total"] or Decimal("0")
    upcoming = Reminder.objects.filter(user=user, completed=False, remind_at__gte=timezone.now()).order_by("remind_at")[:5]
    recent_notes = Note.objects.filter(user=user).order_by("-updated_at")[:5]
    return JsonResponse({
        "counts": {
            "notes": Note.objects.filter(user=user).count(),
            "todos_open": Todo.objects.filter(user=user, completed=False).count(),
            "files": UploadedFile.objects.filter(user=user, is_user_deleted=False).count(),
            "bookmarks": Bookmark.objects.filter(user=user).count(),
            "shortlinks": ShortLink.objects.filter(user=user).count(),
            "reminders_open": Reminder.objects.filter(user=user, completed=False).count(),
        },
        "expense_month_total": str(expense_total),
        "recent_notes": [{"id": n.id, "title": n.title or "无标题", "updated_at": n.updated_at.isoformat()} for n in recent_notes],
        "upcoming_reminders": [{"id": r.id, "title": r.title, "remind_at": r.remind_at.isoformat()} for r in upcoming],
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def text_tool_log_api(request):
    action_name = request.data.get("action", "process")
    detail = request.data.get("detail", "")[:500]
    ToolUsage.objects.create(user=request.user, tool="text", action=action_name, detail=detail)
    return JsonResponse({"ok": True})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def image_process_api(request):
    upload = request.FILES.get("file")
    operation = request.POST.get("operation", "compress")
    if not upload:
        return JsonResponse({"error": "请上传图片"}, status=400)
    try:
        from PIL import Image
        img = Image.open(upload)
        if img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGB")

        if operation == "base64":
            raw = upload.read()
            text = "data:%s;base64,%s" % (upload.content_type or "image/png", base64.b64encode(raw).decode("ascii"))
            rec = ImageProcessHistory.objects.create(
                user=request.user, operation=operation, original_name=upload.name, result_text=text[:100000]
            )
            ToolUsage.objects.create(user=request.user, tool="image", action=operation, detail=upload.name)
            return JsonResponse({"success": True, "id": rec.id, "result_text": text})

        if operation == "target":
            target_kb = int(request.POST.get("target_kb") or 1024)
            fmt = request.POST.get("format", "WEBP").upper()
            data, ext, final_size, final_width, final_height = _compress_image_file_to_target(
                upload,
                target_kb=target_kb,
                fmt=fmt,
            )
            base_name = _os.path.splitext(upload.name)[0]
            result_name = f"{base_name}_under_{target_kb}kb.{ext}"
            rec = ImageProcessHistory.objects.create(
                user=request.user, operation=operation, original_name=upload.name
            )
            rec.result_file.save(result_name, ContentFile(data), save=True)
            ToolUsage.objects.create(user=request.user, tool="image", action=operation, detail=f"{upload.name} <= {target_kb}KB")
            serializer = ImageProcessHistorySerializer(rec, context={"request": request})
            payload = serializer.data
            payload["final_size"] = final_size
            payload["final_width"] = final_width
            payload["final_height"] = final_height
            return JsonResponse({"success": True, "data": payload})

        width = int(request.POST.get("width") or 0)
        height = int(request.POST.get("height") or 0)
        if operation == "resize" and width > 0 and height > 0:
            img = img.resize((width, height))

        fmt = request.POST.get("format", "WEBP").upper()
        if fmt not in ("WEBP", "PNG", "JPEG"):
            fmt = "WEBP"
        quality = max(10, min(95, int(request.POST.get("quality") or 80)))
        ext = "jpg" if fmt == "JPEG" else fmt.lower()
        output = io.BytesIO()
        save_kwargs = {"format": fmt}
        if fmt in ("WEBP", "JPEG"):
            save_kwargs["quality"] = quality
        img.save(output, **save_kwargs)
        output.seek(0)

        base_name = _os.path.splitext(upload.name)[0]
        result_name = f"{base_name}_{operation}.{ext}"
        rec = ImageProcessHistory.objects.create(
            user=request.user, operation=operation, original_name=upload.name
        )
        rec.result_file.save(result_name, ContentFile(output.getvalue()), save=True)
        ToolUsage.objects.create(user=request.user, tool="image", action=operation, detail=upload.name)
        serializer = ImageProcessHistorySerializer(rec, context={"request": request})
        return JsonResponse({"success": True, "data": serializer.data})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)
@api_view(["GET"])
def server_status_api(request):
    """服务器状态（仅管理员可见）"""
    if not request.user.is_authenticated or not request.user.is_superuser:
        return JsonResponse({"error": "无权限"}, status=403)
    try:
        import json, subprocess, shutil
        # 磁盘信息
        disk_total = disk_used = disk_percent = 0
        try:
            du = shutil.disk_usage("/")
            disk_total = round(du.total / (1024**3), 1)
            disk_used = round(du.used / (1024**3), 1)
            disk_percent = round(du.used / du.total * 100, 1)
        except:
            pass
        
        # 内存信息
        mem_total = mem_used = mem_percent = 0
        try:
            r = subprocess.run(["free", "-b"], capture_output=True, text=True, timeout=5)
            for ln in r.stdout.split(chr(10)):
                if ln.startswith("Mem:"):
                    parts = ln.split()
                    if len(parts) >= 3:
                        mem_total = int(parts[1])
                        mem_used = int(parts[2])
                        mem_percent = round(mem_used / mem_total * 100, 1) if mem_total > 0 else 0
        except:
            pass
        
        # CPU 信息（使用 /proc/stat 更可靠）
        cpu_percent = 0
        try:
            with open("/proc/stat") as f:
                line = f.readline()
            parts = line.strip().split()
            if len(parts) >= 5:
                values = [int(x) for x in parts[1:]]
                total = sum(values)
                idle = values[3]
                if total > 0:
                    cpu_percent = round((1 - idle / total) * 100, 1)
        except:
            pass

        # 统计网站访问次数
        try:
            from .models import AccessLog
            visit_count = AccessLog.objects.filter(log_type="visit").count()
            user_count = AccessLog.objects.filter(log_type="visit").values("ip_address").distinct().count()
        except:
            visit_count = 0
            user_count = 0

        
        return JsonResponse({
            "disk": {"total": disk_total, "used": disk_used, "free": round(disk_total - disk_used, 1), "percent": disk_percent},
            "memory": {"total": mem_total, "used": mem_used, "percent": mem_percent},
            "cpu": {"percent": cpu_percent}, "visits": {"total": visit_count, "unique_ips": user_count},
        })
    except Exception as e:
        return JsonResponse({"error": str(e), "detail": "服务器状态获取失败"}, status=500)


@api_view(["POST"])
def deploy_update_api(request):
    if not request.user.is_authenticated or not request.user.is_superuser:
        return JsonResponse({"error": "无权限"}, status=403)
    try:
        from django.core.management import call_command
        from django.core.management.base import CommandError
        from io import StringIO
        out = StringIO()
        err = StringIO()
        try:
            call_command("deploy_update", stdout=out, stderr=err)
            success = True
            msg = "更新完成"
        except CommandError as e:
            success = False
            msg = str(e)
        except SystemExit as e:
            success = False
            msg = f"更新失败 (code {e.code})"
        return JsonResponse({
            "success": success, "message": msg,
            "output": out.getvalue(), "error": err.getvalue()
        })
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)




# ===== 文件转换处理 =====
import os as _os
import csv
import json as _json
from io import BytesIO
from django.core.files.base import ContentFile
from django.views.decorators.csrf import csrf_exempt

def _get_file_path(relative_path):
    return _os.path.join(settings.MEDIA_ROOT, relative_path)

@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def file_convert_api(request):
    """文件转换统一入口"""
    if request.method != "POST":
        return JsonResponse({"error": "\u4ec5\u652f\u6301POST"}, status=405)
    conv_type = request.POST.get("conv_type", "")
    upload = request.FILES.get("file")
    if not conv_type:
        return JsonResponse({"error": "\u8bf7\u6307\u5b9a\u8f6c\u6362\u7c7b\u578b"}, status=400)
    if not upload:
        return JsonResponse({"error": "\u8bf7\u4e0a\u4f20\u6587\u4ef6"}, status=400)
    try:
        rec = FileConversion(user=request.user)
        rec.conv_type = conv_type
        rec.original_name = upload.name
        rec.status = "processing"
        rec.source_file.save(upload.name, upload, save=True)
        if conv_type == "pdf2word":
            result = _convert_pdf_to_word(rec)
        elif conv_type == "pdf2txt":
            result = _convert_pdf_to_text(rec)
        elif conv_type == "pdf2images":
            result = _convert_pdf_to_images(rec)
        elif conv_type == "merge_pdf":
            result = _merge_pdfs([rec])
        elif conv_type == "split_pdf":
            page = request.POST.get("page", "1")
            result = _split_pdf(rec, page)
        elif conv_type == "word2pdf":
            result = _convert_word_to_pdf(rec)
        elif conv_type == "excel2csv":
            result = _convert_excel_to_csv(rec)
        elif conv_type == "excel2json":
            result = _convert_excel_to_json(rec)
        else:
            return JsonResponse({"error": f"\u4e0d\u652f\u6301\u7684\u8f6c\u6362\u7c7b\u578b: {conv_type}"}, status=400)
        rec.status = "completed"
        if result.get("path"):
            rec.result_file = result["path"]
        rec.save()
        ToolUsage.objects.create(user=request.user, tool="file", action=conv_type, detail=f"\u6587\u4ef6\u8f6c\u6362: {conv_type}")
        return JsonResponse({"success": True, "message": result.get("message", "\u8f6c\u6362\u6210\u529f"), "filename": result.get("filename", ""), "download_url": result.get("download_url", "")})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

def _convert_pdf_to_word(rec):
    from pdf2docx import Converter
    input_path = rec.source_file.path
    base = _os.path.splitext(_os.path.basename(input_path))[0]
    output_name = f"{base}.docx"
    output_rel = f"conversions/output/{output_name}"
    output_path = _os.path.join(settings.MEDIA_ROOT, output_rel)
    _os.makedirs(_os.path.dirname(output_path), exist_ok=True)
    cv = Converter(input_path)
    cv.convert(output_path, start=0, end=None)
    cv.close()
    return {"message": "PDF \u5df2\u8f6c\u6362\u4e3a Word", "filename": output_name, "path": output_rel, "download_url": f"{settings.MEDIA_URL}{output_rel}"}

def _convert_pdf_to_text(rec):
    import fitz
    input_path = rec.source_file.path
    doc = fitz.open(input_path)
    text = ""
    for page in doc: text += page.get_text() + "\n"
    doc.close()
    base = _os.path.splitext(_os.path.basename(input_path))[0]
    output_name = f"{base}.txt"
    output_rel = f"conversions/output/{output_name}"
    output_path = _os.path.join(settings.MEDIA_ROOT, output_rel)
    _os.makedirs(_os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f: f.write(text)
    return {"message": f"\u5df2\u63d0\u53d6\u6587\u672c ({len(text)} \u5b57\u7b26)", "filename": output_name, "path": output_rel, "download_url": f"{settings.MEDIA_URL}{output_rel}"}

def _convert_pdf_to_images(rec):
    import fitz; import zipfile
    input_path = rec.source_file.path
    doc = fitz.open(input_path)
    base = _os.path.splitext(_os.path.basename(input_path))[0]
    output_dir_rel = f"conversions/output/{base}_images"
    output_dir = _os.path.join(settings.MEDIA_ROOT, output_dir_rel)
    _os.makedirs(output_dir, exist_ok=True)
    for i, page in enumerate(doc):
        pix = page.get_pixmap(dpi=150)
        pix.save(_os.path.join(output_dir, f"{base}_page_{i+1}.png"))
    doc.close()
    zip_name = f"{base}_images.zip"
    zip_rel = f"conversions/output/{zip_name}"
    zip_path = _os.path.join(settings.MEDIA_ROOT, zip_rel)
    with zipfile.ZipFile(zip_path, "w") as zf:
        for fn in _os.listdir(output_dir):
            zf.write(_os.path.join(output_dir, fn), fn)
    return {"message": f"\u5df2\u751f\u6210 {len(doc)} \u5f20\u56fe\u7247", "filename": zip_name, "path": zip_rel, "download_url": f"{settings.MEDIA_URL}{zip_rel}"}

def _merge_pdfs(recs):
    import fitz
    merged = fitz.open()
    for rec in recs:
        fp = rec.source_file.path if hasattr(rec, "source_file") else rec.file.path
        if fp.lower().endswith(".pdf"):
            doc = fitz.open(fp); merged.insert_pdf(doc); doc.close()
    output_name = f"merged_{_os.urandom(4).hex()}.pdf"
    output_rel = f"conversions/output/{output_name}"
    output_path = _os.path.join(settings.MEDIA_ROOT, output_rel)
    _os.makedirs(_os.path.dirname(output_path), exist_ok=True)
    merged.save(output_path); merged.close()
    return {"message": f"\u5df2\u5408\u5e76 {len(recs)} \u4e2a PDF", "filename": output_name, "path": output_rel, "download_url": f"{settings.MEDIA_URL}{output_rel}"}

def _split_pdf(rec, page_str):
    import fitz; import zipfile
    input_path = rec.source_file.path
    doc = fitz.open(input_path)
    total = len(doc)
    pages = [int(p.strip()) for p in page_str.split(",") if p.strip().isdigit()]
    base = _os.path.splitext(_os.path.basename(input_path))[0]
    output_dir_rel = f"conversions/output/{base}_split"
    output_dir = _os.path.join(settings.MEDIA_ROOT, output_dir_rel)
    _os.makedirs(output_dir, exist_ok=True)
    results = []
    for p in pages:
        if 1 <= p <= total:
            new_doc = fitz.open()
            new_doc.insert_pdf(doc, from_page=p-1, to_page=p-1)
            name = f"{base}_page_{p}.pdf"
            new_doc.save(_os.path.join(output_dir, name)); new_doc.close()
            results.append(name)
    doc.close()
    zip_name = f"{base}_split_pages.zip"
    zip_rel = f"conversions/output/{zip_name}"
    zip_path = _os.path.join(settings.MEDIA_ROOT, zip_rel)
    with zipfile.ZipFile(zip_path, "w") as zf:
        for name in results: zf.write(_os.path.join(output_dir, name), name)
    return {"message": f"\u5df2\u62c6\u5206 {len(results)} \u9875", "filename": zip_name, "path": zip_rel, "download_url": f"{settings.MEDIA_URL}{zip_rel}"}

def _convert_word_to_pdf(rec):
    from docx import Document; import fitz
    input_path = rec.source_file.path
    base = _os.path.splitext(_os.path.basename(input_path))[0]
    doc = Document(input_path)
    pdf_doc = fitz.open()
    page = pdf_doc.new_page()
    page.insert_text(fitz.Point(50, 50), "\n".join([p.text for p in doc.paragraphs]), fontsize=11)
    output_name = f"{base}.pdf"
    output_rel = f"conversions/output/{output_name}"
    output_path = _os.path.join(settings.MEDIA_ROOT, output_rel)
    _os.makedirs(_os.path.dirname(output_path), exist_ok=True)
    pdf_doc.save(output_path); pdf_doc.close()
    return {"message": "Word \u5df2\u8f6c\u6362\u4e3a PDF", "filename": output_name, "path": output_rel, "download_url": f"{settings.MEDIA_URL}{output_rel}"}

def _convert_excel_to_csv(rec):
    import openpyxl; import csv
    input_path = rec.source_file.path
    base = _os.path.splitext(_os.path.basename(input_path))[0]
    wb = openpyxl.load_workbook(input_path, read_only=True)
    ws = wb.active
    output_name = f"{base}.csv"
    output_rel = f"conversions/output/{output_name}"
    output_path = _os.path.join(settings.MEDIA_ROOT, output_rel)
    _os.makedirs(_os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(list(ws.iter_rows(values_only=True)))
    wb.close()
    return {"message": "Excel \u5df2\u8f6c\u6362\u4e3a CSV", "filename": output_name, "path": output_rel, "download_url": f"{settings.MEDIA_URL}{output_rel}"}

def _convert_excel_to_json(rec):
    import openpyxl; import json
    input_path = rec.source_file.path
    wb = openpyxl.load_workbook(input_path, read_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(next(ws.iter_rows(values_only=True)))]
        result[sheet_name] = [{headers[i]: (v if v is not None else "") for i, v in enumerate(row)} for row in ws.iter_rows(values_only=True)]
    wb.close()
    base = _os.path.splitext(_os.path.basename(input_path))[0]
    output_name = f"{base}.json"
    output_rel = f"conversions/output/{output_name}"
    output_path = _os.path.join(settings.MEDIA_ROOT, output_rel)
    _os.makedirs(_os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f: json.dump(result, f, ensure_ascii=False, indent=2)
    return {"message": "Excel \u5df2\u8f6c\u6362\u4e3a JSON", "filename": output_name, "path": output_rel, "download_url": f"{settings.MEDIA_URL}{output_rel}"}

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def file_conversion_list(request):
    """获取用户的转换记录"""
    if not request.user.is_authenticated:
        return JsonResponse({"error": "\u672a\u767b\u5f55"}, status=401)
    qs = FileConversion.objects.filter(user=request.user).order_by("-created_at")[:20]
    data = []
    for c in qs:
        dl_url = None
        try:
            if c.result_file and _os.path.exists(c.result_file.path):
                dl_url = c.result_file.url
        except: pass
        data.append({"id": c.id, "conv_type": c.conv_type, "conv_type_display": c.get_conv_type_display(), "original_name": c.original_name, "status": c.status, "download_url": dl_url, "error_msg": c.error_msg, "created_at": c.created_at.isoformat()})
    return JsonResponse({"data": data})
